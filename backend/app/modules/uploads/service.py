import csv
import io
from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal

from fastapi import UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.db.models import Portfolio, UploadJob
from app.db.models import User
from app.modules.portfolios.service import PortfolioService
from app.modules.uploads.errors import UploadJobNotFoundError, UploadValidationError
from app.modules.uploads.repository import (
    UploadRepository,
    job_mapping,
    row_errors,
    row_mapped_data,
    row_raw_data,
)
from app.modules.uploads.schemas import (
    ColumnMappingResponse,
    ConfirmUploadResponse,
    REQUIRED_UPLOAD_FIELDS,
    SUPPORTED_UPLOAD_FIELDS,
    UploadErrorsResponse,
    UploadJobResponse,
    UploadRowResponse,
    ValidateUploadResponse,
)


class UploadService:
    def __init__(self, db: Session):
        self.repository = UploadRepository(db)
        self.portfolio_service = PortfolioService(db)

    async def create_upload(self, *, portfolio_id: str, user: User, file: UploadFile) -> UploadJobResponse:
        portfolio = self.portfolio_service.get_portfolio(portfolio_id=portfolio_id, user=user)
        filename = file.filename or "upload"
        content = await file.read()
        rows = parse_upload_rows(filename=filename, content=content)
        if not rows:
            raise UploadValidationError("Upload file contains no data rows")

        upload_job = self.repository.create_job(portfolio=portfolio, filename=filename, rows=rows)
        return self.to_job_response(upload_job)

    def get_upload(self, *, upload_job_id: str, user: User) -> UploadJobResponse:
        upload_job, _portfolio = self.get_job_for_user(upload_job_id=upload_job_id, user=user)
        return self.to_job_response(upload_job)

    def apply_column_mapping(
        self,
        *,
        upload_job_id: str,
        user: User,
        mapping: dict[str, str],
    ) -> ColumnMappingResponse:
        upload_job, _portfolio = self.get_job_for_user(upload_job_id=upload_job_id, user=user)
        normalized_mapping = normalize_mapping(mapping)
        detected_columns = detected_columns_for_job(upload_job)
        unknown_fields = sorted(set(normalized_mapping) - SUPPORTED_UPLOAD_FIELDS)
        if unknown_fields:
            raise UploadValidationError(f"Unsupported mapped fields: {', '.join(unknown_fields)}")

        missing_required = sorted(REQUIRED_UPLOAD_FIELDS - set(normalized_mapping))
        if missing_required:
            raise UploadValidationError(f"Missing required mapping fields: {', '.join(missing_required)}")

        missing_columns = sorted(set(normalized_mapping.values()) - set(detected_columns))
        if missing_columns:
            raise UploadValidationError(f"Mapped source columns not found: {', '.join(missing_columns)}")

        upload_job = self.repository.save_mapping(upload_job=upload_job, mapping=normalized_mapping)
        mapped_rows = {row.id: map_row(row_raw_data(row), normalized_mapping) for row in upload_job.rows}
        upload_job = self.repository.save_mapped_rows(upload_job=upload_job, mapped_rows=mapped_rows)
        return ColumnMappingResponse(
            upload_job=self.to_job_response(upload_job),
            mapping=normalized_mapping,
            mapped_preview_rows=[row_mapped_data(row) for row in upload_job.rows[:5]],
        )

    def validate_upload(self, *, upload_job_id: str, user: User) -> ValidateUploadResponse:
        upload_job, portfolio = self.get_job_for_user(upload_job_id=upload_job_id, user=user)
        mapping = job_mapping(upload_job)
        if not mapping:
            raise UploadValidationError("Column mapping is required before validation")

        row_results = {}
        for row in upload_job.rows:
            mapped = row_mapped_data(row) or map_row(row_raw_data(row), mapping)
            normalized, errors = validate_mapped_row(mapped=mapped, portfolio=portfolio)
            row_results[row.id] = ("invalid" if errors else "valid", normalized, errors)

        upload_job = self.repository.save_validation(upload_job=upload_job, row_results=row_results)
        return ValidateUploadResponse(
            upload_job=self.to_job_response(upload_job),
            rows=[self.to_row_response(row) for row in upload_job.rows],
        )

    def confirm_upload(self, *, upload_job_id: str, user: User) -> ConfirmUploadResponse:
        upload_job, portfolio = self.get_job_for_user(upload_job_id=upload_job_id, user=user)
        if upload_job.status != "validated":
            raise UploadValidationError("Upload must be validated before confirm import")

        existing_symbols = self.repository.existing_symbols(portfolio=portfolio)
        symbols_seen = set(existing_symbols)
        rows_to_import: list[dict] = []
        warnings: list[str] = []
        skipped_count = 0

        for row in upload_job.rows:
            if row.status != "valid":
                continue
            mapped = row_mapped_data(row)
            symbol = mapped["symbol"]
            if symbol in symbols_seen:
                skipped_count += 1
                warnings.append(
                    f"Row {row.row_number}: {symbol} skipped because it already exists in the portfolio or upload batch"
                )
                continue
            symbols_seen.add(symbol)
            rows_to_import.append(mapped)

        holdings = self.repository.create_holdings(portfolio=portfolio, rows=rows_to_import)
        if upload_job.invalid_rows or skipped_count:
            status = "partial_imported"
            if upload_job.invalid_rows:
                warnings.append(
                    f"Partial import occurred: {upload_job.invalid_rows} invalid row(s) were not imported"
                )
        else:
            status = "imported"

        self.repository.mark_imported(upload_job=upload_job, status=status)
        return ConfirmUploadResponse(
            upload_job_id=upload_job.id,
            status=status,
            imported_count=len(holdings),
            skipped_count=skipped_count,
            invalid_rows=upload_job.invalid_rows,
            warnings=warnings,
            created_holding_ids=[holding.id for holding in holdings],
        )

    def get_errors(self, *, upload_job_id: str, user: User) -> UploadErrorsResponse:
        upload_job, _portfolio = self.get_job_for_user(upload_job_id=upload_job_id, user=user)
        return UploadErrorsResponse(
            upload_job_id=upload_job.id,
            errors=[self.to_row_response(row) for row in upload_job.rows if row.status == "invalid"],
        )

    def get_job_for_user(self, *, upload_job_id: str, user: User) -> tuple[UploadJob, Portfolio]:
        upload_job = self.repository.get_job(upload_job_id=upload_job_id)
        if upload_job is None:
            raise UploadJobNotFoundError()
        portfolio = self.portfolio_service.get_portfolio(portfolio_id=upload_job.portfolio_id, user=user)
        return upload_job, portfolio

    @staticmethod
    def to_job_response(upload_job: UploadJob) -> UploadJobResponse:
        rows = upload_job.rows
        return UploadJobResponse(
            id=upload_job.id,
            portfolio_id=upload_job.portfolio_id,
            filename=upload_job.filename,
            status=upload_job.status,
            total_rows=upload_job.total_rows,
            valid_rows=upload_job.valid_rows,
            invalid_rows=upload_job.invalid_rows,
            detected_columns=detected_columns_for_job(upload_job),
            preview_rows=[row_raw_data(row) for row in rows[:5]],
            created_at=upload_job.created_at,
            updated_at=upload_job.updated_at,
        )

    @staticmethod
    def to_row_response(row) -> UploadRowResponse:
        return UploadRowResponse(
            id=row.id,
            row_number=row.row_number,
            raw_data=row_raw_data(row),
            mapped_data=row_mapped_data(row),
            validation_errors=row_errors(row),
            status=row.status,
        )


def parse_upload_rows(*, filename: str, content: bytes) -> list[dict]:
    lower_filename = filename.lower()
    if lower_filename.endswith(".csv"):
        return parse_csv(content)
    if lower_filename.endswith(".xlsx"):
        return parse_xlsx(content)
    raise UploadValidationError("Only CSV and XLSX uploads are supported")


def parse_csv(content: bytes) -> list[dict]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    return [normalize_raw_row(row) for row in reader]


def parse_xlsx(content: bytes) -> list[dict]:
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    output: list[dict] = []
    for row in rows[1:]:
        raw = OrderedDict()
        for header, value in zip(headers, row):
            if header:
                raw[header] = json_safe(value)
        output.append(dict(raw))
    return output


def normalize_raw_row(row: dict) -> dict:
    return {str(key).strip(): json_safe(value) for key, value in row.items() if key is not None}


def json_safe(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    return value


def detected_columns_for_job(upload_job: UploadJob) -> list[str]:
    columns: list[str] = []
    seen: set[str] = set()
    for row in upload_job.rows:
        for column in row_raw_data(row):
            if column not in seen:
                seen.add(column)
                columns.append(column)
    return columns


def normalize_mapping(mapping: dict[str, str]) -> dict[str, str]:
    normalized: dict[str, str] = {}
    for target_field, source_column in mapping.items():
        field = target_field.strip()
        column = source_column.strip()
        if field and column:
            normalized[field] = column
    return normalized


def map_row(raw_row: dict, mapping: dict[str, str]) -> dict:
    return {target_field: raw_row.get(source_column) for target_field, source_column in mapping.items()}


def validate_mapped_row(*, mapped: dict, portfolio: Portfolio) -> tuple[dict, list[str]]:
    errors: list[str] = []
    normalized: dict = {}

    symbol = clean_string(mapped.get("symbol"))
    if not symbol:
        errors.append("symbol is required")
    else:
        normalized["symbol"] = symbol.upper()

    quantity = parse_optional_number(mapped.get("quantity"))
    if quantity is None:
        errors.append("quantity is required")
    elif quantity <= 0:
        errors.append("quantity must be a positive number")
    else:
        normalized["quantity"] = quantity

    average_cost = parse_optional_number(mapped.get("average_cost"))
    if average_cost is not None and average_cost < 0:
        errors.append("average_cost cannot be negative")
    normalized["average_cost"] = average_cost

    market_value = parse_optional_number(mapped.get("market_value"))
    if market_value is not None and market_value < 0:
        errors.append("market_value cannot be negative")
    normalized["market_value"] = market_value

    if quantity and quantity > 0 and market_value is not None:
        normalized["current_price"] = market_value / quantity
    else:
        normalized["current_price"] = None

    normalized["company_name"] = clean_string(mapped.get("company_name"))
    normalized["currency"] = (clean_string(mapped.get("currency")) or portfolio.base_currency).upper()
    normalized["sector"] = clean_string(mapped.get("sector"))
    normalized["asset_class"] = clean_string(mapped.get("asset_class")) or "Equity"
    normalized["exchange"] = clean_string(mapped.get("exchange"))

    return normalized, errors


def parse_optional_number(value) -> float | None:
    if value is None:
        return None
    if isinstance(value, str) and not value.strip():
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def clean_string(value) -> str | None:
    if value is None:
        return None
    cleaned = str(value).strip()
    return cleaned or None
